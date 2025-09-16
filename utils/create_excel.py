import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import datetime

# Создаем данные
data = {
    'Задача': ['сдача КТС', 'сдача КТС', 'сдача КТС'],
    'Ссылка на Jira': [
        'https://jira.example.com/TASK-123',
        'https://jira.example.com/TASK-124', 
        'https://jira.example.com/TASK-125'
    ],
    'Ссылка на АРМ ЕФС': [
        'https://arm-efs.example.com/system1',
        'https://arm-efs.example.com/system2',
        'https://arm-efs.example.com/system3'
    ],
    'Ссылка на OpenShift (блок 1)': [
        'https://openshift.example.com/block1/app1',
        'https://openshift.example.com/block1/app2',
        'https://openshift.example.com/block1/app3'
    ]
}

# Создаем DataFrame
df = pd.DataFrame(data)

# Создаем Excel файл
filename = f'задачи_КТС_{datetime.datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'

with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Задачи КТС', index=False)
    
    # Получаем workbook и worksheet для форматирования
    workbook = writer.book
    worksheet = writer.sheets['Задачи КТС']
    
    # Форматируем заголовки
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, value in enumerate(df.columns.values, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Автоподбор ширины колонок
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

print(f'Файл "{filename}" успешно создан!')
